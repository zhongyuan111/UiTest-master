#!/usr/bin/env python3
# -*- coding:utf-8 -*-
import os
from openpyxl import load_workbook
import pandas as pd
from config.conf import EXCEL_PATH
import xlrd
import numpy

class ParseExcel(object):
    '''解析excel文件'''
    def __init__(self):
        self.wk = load_workbook(EXCEL_PATH)
        self.excelFile = EXCEL_PATH
    def getSheetByName(self, sheetName):
         """获取sheet对象"""
         sheet = self.wk[sheetName]
         return sheet

    def getRowNum(self, sheet):
         """获取有效数据的最大行号"""
         return sheet.max_row

    def getColsNum(self, sheet):
        """获取有效数据的最大列号"""
        return sheet.max_column

    def getRowValues(self, sheet, rowNum):
        """获取某一行的数据"""
        maxColsNum = self.getColsNum(sheet)
        rowValues = []
        for colsNum in range(1, maxColsNum + 1):
            value = sheet.cell(rowNum, colsNum).value
            if value is None:
                value = ''
            rowValues.append(value)
        return tuple(rowValues)

    def getColumnValues(self, sheet, columnNum):

        """获取某一列的数据"""
        maxRowNum = self.getRowNum(sheet)
        columnValues = []
        for rowNum in range(2, maxRowNum + 1):
            value = sheet.cell(rowNum, columnNum).value
            if value is None:
                value = ''
            columnValues.append(value)
        return tuple(columnValues)

    def getValueOfCell(self, sheet, rowNum, columnNum):
        """获取某一个单元格的数据"""
        value = sheet.cell(rowNum, columnNum).value
        if value is None:
            value = ''
        return value

    def getAllValuesOfSheet(self, sheet):
        """获取某一个sheet页的所有测试数据，返回一个元祖组成的列表"""
        maxRowNum = self.getRowNum(sheet)
        columnNum = self.getColsNum(sheet)
        allValues = []
        for row in range(2, maxRowNum + 1):
            rowValues = []
            for column in range(1, columnNum + 1):
                value = sheet.cell(row, column).value
                if value is None:
                    value = ''
                rowValues.append(value)
            allValues.append(tuple(rowValues))
        return allValues

class ParsePdExcel():
    def __init__(self):
        self.excelFile = EXCEL_PATH
    def pd_read_excel(self,sheet):
        st_data = pd.read_excel(self.excelFile,sheet)
        #print("*************", st_data.values)
        return st_data.values

if __name__ == '__main__':
    excel = ParseExcel()
    pdexcel = ParsePdExcel()
    sheet = excel.getSheetByName('phone')
    print('行号:', excel.getRowNum(sheet))
    print('列号:', excel.getColsNum(sheet))

    allvalues = pdexcel.pd_read_excel('phone')
    allvalues1 = excel.getAllValuesOfSheet(sheet)

    print(allvalues1)
    print(allvalues)