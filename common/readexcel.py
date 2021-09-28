import os
from openpyxl import load_workbook
import pandas as pd
import xlrd
import numpy
from config.conf import EXCEL_PATH



# 读取Yaml文件和Json文件
# def test_read_data_from_json_yaml(data_file):
#     return_value = []
#     data_file_path = os.path.abspath(data_file)
#     print(data_file_path)
#
#     _is_yaml_file = data_file_path.endswith((".yml", ".yaml"))
#     with codecs.open(data_file_path, 'r', 'utf-8') as f:
#         # 从YAML或JSON文件中加载数据
#         if _is_yaml_file:
#             data = yaml.safe_load(f)
#         else:
#             data = json.load(f)
#
#     for i, elem in enumerate(data):
#         if isinstance(data, dict):
#             key, value = elem, data[elem]
#             if isinstance(value, dict):
#                 case_data = []
#                 for v in value.values():
#                     case_data.append(v)
#                 return_value.append(tuple(case_data))
#             else:
#                 return_value.append((value,))
#     return return_value

def test_read_data_from_excel(sheet_name):
    return_value = []

    # 判断文件是否存在
    if not os.path.exists(excel_file):
        raise ValueError("File not exists")

    # 打开指定的sheet
    wb = load_workbook(excel_file)

    # 按照pytest接受的格式输出数据
    for s in wb.sheetnames:
        if s == sheet_name:
            sheet = wb[sheet_name]
            for row in sheet.rows:
                value = [col.value for col in row]
                return_value.append([col.value for col in row])
    #print("*********测试1*********", return_value)
    print ("*********测试*********",return_value[1:])
    #print("*********测试4*********",list(b))

# 读取Excel文件 -- Pandas
def test_read_data_from_pandas(sheet_name):
    if not os.path.exists(excel_file):
        raise ValueError("File not exists")
    s = pd.ExcelFile(excel_file)
    df = s.parse(sheet_name)

    #return df.values.tolist()

    nested_lst_of_tuples = [tuple(l) for l in df.values.tolist()]


    print("*********测试3**&&&&&&*******", df.values.tolist())
    # number = df.values.tolist()[0]
    # str(number.replace("'", ""))
    # print("*********测试6**&&&&&&*******",number)

def pd_read_excel(sheet_name):
    if not os.path.exists(excel_file):
        raise ValueError("File not exists")
    st_data = pd.read_excel(excel_file, sheet_name=sheet_name)
    # 2.读取book表所有数据(list形式)
    print("*************",st_data.values)
    return st_data.values


# test_read_data_from_excel('login')
# test_read_data_from_pandas('login')
# test_pd_read_excel('login')








